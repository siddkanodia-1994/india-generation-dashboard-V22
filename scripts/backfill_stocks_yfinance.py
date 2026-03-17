"""
Backfill historical stock closing prices into stock.xlsx (Prices sheet)
using yfinance for NSE-listed tickers.

Usage:
  python backfill_stocks_yfinance.py --start-date 2026-01-20 --end-date 2026-03-16

Defaults:
  --start-date: day after the last row in stock.xlsx that has price data
  --end-date:   yesterday (IST)
"""

import argparse
import sys
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

import pandas as pd
import yfinance as yf
import openpyxl

sys.path.insert(0, str(Path(__file__).parent))
from config import CSV_PATHS, STOCK_TICKERS

IST = timezone(timedelta(hours=5, minutes=30))


# ── Helpers ───────────────────────────────────────────────────────────────────

def _parse_date(val):
    import datetime as dt
    if isinstance(val, dt.datetime):
        return val.date()
    if isinstance(val, dt.date):
        return val
    if isinstance(val, str):
        try:
            parts = val.strip().split("/")
            d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
            if y < 100:
                y += 2000
            return date(y, m, d)
        except Exception:
            pass
    return None


def _get_column_index(ws, ticker: str):
    for col in range(1, ws.max_column + 2):
        val = ws.cell(row=1, column=col).value
        if val is None:
            return None
        if str(val).strip().upper() == ticker.upper():
            return col
    return None


def _ensure_ticker_column(ws, ticker: str) -> int:
    idx = _get_column_index(ws, ticker)
    if idx is not None:
        return idx
    new_col = ws.max_column + 1
    ws.cell(row=1, column=new_col, value=ticker)
    return new_col


def _last_filled_date(ws):
    """Return the last date in col A where col 2 (first ticker) has actual data."""
    last = None
    for row in ws.iter_rows(min_row=2, max_col=2, values_only=True):
        date_val, price_val = row[0], row[1]
        if date_val is None:
            continue
        if price_val in (None, ""):
            continue
        d = _parse_date(date_val)
        if d:
            last = d
    return last


# ── Core backfill ─────────────────────────────────────────────────────────────

def backfill(start: date, end: date) -> None:
    xlsx_path = CSV_PATHS["stocks"]
    wb = openpyxl.load_workbook(xlsx_path)
    prices_ws = wb["Prices"] if "Prices" in wb.sheetnames else wb.active

    # Build date → row index from column A
    date_row = {}
    for r in range(2, prices_ws.max_row + 1):
        val = prices_ws.cell(r, 1).value
        if val:
            d = _parse_date(val)
            if d:
                date_row[d] = r

    print(f"[BACKFILL] Date range: {start} → {end}")
    print(f"[BACKFILL] Tickers: {STOCK_TICKERS}")

    symbols = [f"{t}.NS" for t in STOCK_TICKERS]
    print(f"[BACKFILL] Downloading yfinance data for {len(symbols)} symbols...")

    raw = yf.download(
        symbols,
        start=start,
        end=end + timedelta(days=1),
        auto_adjust=True,
        progress=False,
    )

    # yf.download returns MultiIndex columns when multiple tickers;
    # select "Close" level
    if isinstance(raw.columns, pd.MultiIndex):
        close_data = raw["Close"]
    else:
        # Single ticker — columns is just ["Close"]
        close_data = raw[["Close"]]
        close_data.columns = symbols

    rows_written = 0
    for dt_idx, row_data in close_data.iterrows():
        d = dt_idx.date() if hasattr(dt_idx, "date") else dt_idx
        if d < start or d > end:
            continue

        if d not in date_row:
            new_r = prices_ws.max_row + 1
            prices_ws.cell(new_r, 1, d.strftime("%d/%m/%y"))
            date_row[d] = new_r
            print(f"[BACKFILL] Added new row for {d} at row {new_r}")

        r = date_row[d]
        any_written = False
        for ticker in STOCK_TICKERS:
            sym = f"{ticker}.NS"
            price = row_data.get(sym)
            if price is not None and not pd.isna(price):
                col = _ensure_ticker_column(prices_ws, ticker)
                prices_ws.cell(r, col, round(float(price), 2))
                any_written = True
        if any_written:
            rows_written += 1

    wb.save(xlsx_path)
    print(f"[BACKFILL] Done. Wrote prices for {rows_written} trading days ({start} → {end}).")


# ── CLI ───────────────────────────────────────────────────────────────────────

def main():
    parser = argparse.ArgumentParser(description="Backfill stock prices via yfinance")
    parser.add_argument("--start-date", help="YYYY-MM-DD (default: day after last filled row)")
    parser.add_argument("--end-date",   help="YYYY-MM-DD (default: yesterday IST)")
    args = parser.parse_args()

    yesterday = datetime.now(IST).date() - timedelta(days=1)

    if args.end_date:
        end = date.fromisoformat(args.end_date)
    else:
        end = yesterday

    if args.start_date:
        start = date.fromisoformat(args.start_date)
    else:
        # Determine from last filled row
        xlsx_path = CSV_PATHS["stocks"]
        if Path(xlsx_path).exists():
            wb = openpyxl.load_workbook(xlsx_path)
            prices_ws = wb["Prices"] if "Prices" in wb.sheetnames else wb.active
            last = _last_filled_date(prices_ws)
            start = (last + timedelta(days=1)) if last else date(2026, 1, 20)
            print(f"[BACKFILL] Auto-detected start: {start} (last filled: {last})")
        else:
            start = date(2026, 1, 20)

    if start > end:
        print(f"[BACKFILL] Nothing to do: start={start} > end={end}")
        sys.exit(0)

    backfill(start, end)


if __name__ == "__main__":
    main()
