"""
One-time backfill: overwrite 'Stock P-B' sheet rows for Mar 16 – Apr 28, 2026
with consolidated P/B from Trendlyne (PBV_A_SHARE_NOW), replacing Screener.in
standalone values written during that window.
"""

import sys
import time
from datetime import date, datetime, timedelta
from pathlib import Path

import requests

sys.path.insert(0, str(Path(__file__).parent))
from config import CSV_PATHS, STOCK_TICKERS, TICKER_COLUMN_NAMES, TRENDLYNE_IDS

START_DATE = date(2026, 3, 16)
END_DATE   = date(2026, 4, 28)

SESSION = requests.Session()
SESSION.headers.update({
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
})


def _parse_xlsx_date(val):
    """Parse a cell value from 'Stock P-B' column A → date or None."""
    import datetime as dt
    if val is None:
        return None
    if isinstance(val, dt.datetime):
        return val.date()
    if isinstance(val, dt.date):
        return val
    if isinstance(val, str):
        try:
            parts = val.strip().split("/")
            d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
            if y < 100:
                y += 2000
            return date(y, m, d)
        except Exception:
            pass
    return None


def _format_date(d: date) -> str:
    return d.strftime("%d/%m/%y")


def fetch_trendlyne(ticker: str) -> dict[str, float]:
    """Return {date_str(YYYY-MM-DD): pb_value} for START_DATE..END_DATE."""
    tl_id = TRENDLYNE_IDS[ticker]
    try:
        r = SESSION.get(
            f"https://trendlyne.com/mapp/v1/stock/chart-data/{tl_id}/PBV_A_SHARE_NOW/",
            timeout=15,
        )
        r.raise_for_status()
        pts = r.json().get("body", {}).get("eodData", [])
    except Exception as e:
        print(f"  [ERROR] Trendlyne fetch failed for {ticker}: {e}")
        return {}

    result = {}
    for ts_ms, pb in pts:
        d = datetime.fromtimestamp(ts_ms / 1000).date()
        if START_DATE <= d <= END_DATE:
            result[d] = round(pb, 2)
    return result


def main():
    import openpyxl

    xlsx_path = CSV_PATHS["stocks"]
    print(f"Loading {xlsx_path} ...")
    wb = openpyxl.load_workbook(xlsx_path)
    pb_ws = wb["Stock P-B"]

    # Build header map: col_name -> 1-based column index
    headers = {}
    for col in range(1, pb_ws.max_column + 1):
        val = pb_ws.cell(row=1, column=col).value
        if val is not None:
            headers[str(val).strip()] = col

    # Fetch Trendlyne data for all tickers
    # tl_data: {trading_date -> {col_name -> pb_value}}
    tl_data: dict[date, dict[str, float]] = {}
    for ticker in STOCK_TICKERS:
        col_name = TICKER_COLUMN_NAMES.get(ticker, ticker)
        print(f"Fetching Trendlyne P/B for {ticker} ({col_name})...")
        series = fetch_trendlyne(ticker)
        print(f"  Got {len(series)} trading days in range")
        for d, pb in series.items():
            tl_data.setdefault(d, {})[col_name] = pb
        time.sleep(0.5)

    print(f"\nTrendlyne data covers {len(tl_data)} unique dates in range.")

    # Pass 1: overwrite existing rows
    dates_updated: set[date] = set()
    for row_idx in range(2, pb_ws.max_row + 1):
        row_date = _parse_xlsx_date(pb_ws.cell(row_idx, 1).value)
        if row_date is None:
            continue
        if row_date not in tl_data:
            continue
        vals = tl_data[row_date]
        for col_name, pb in vals.items():
            col_idx = headers.get(col_name)
            if col_idx is not None:
                pb_ws.cell(row_idx, col_idx, pb)
        dates_updated.add(row_date)
        print(f"  Overwriting row {row_idx}: {row_date} ({len(vals)} stocks)")

    # Pass 2: insert missing dates (append at bottom; xlsx rows are date-ordered)
    missing = sorted(d for d in tl_data if d not in dates_updated)
    if missing:
        print(f"\nInserting {len(missing)} missing date rows...")
        for d in missing:
            new_row = pb_ws.max_row + 1
            pb_ws.cell(new_row, 1, _format_date(d))
            vals = tl_data[d]
            for col_name, pb in vals.items():
                col_idx = headers.get(col_name)
                if col_idx is not None:
                    pb_ws.cell(new_row, col_idx, pb)
            print(f"  Inserted row {new_row}: {d}")
    else:
        print("\nNo missing dates to insert.")

    wb.save(xlsx_path)
    print(f"\nDone. Updated {len(dates_updated)} rows, inserted {len(missing)} rows.")
    print(f"Saved to {xlsx_path}")


if __name__ == "__main__":
    main()
