"""
Scrape daily stock price and P/B ratio for each ticker from Screener.in
and append to stock.xlsx (sheets: "Prices" and "P/B").

Reads STOCK_TICKERS from config.py.
"""

import re
import sys
import time
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

import requests
from bs4 import BeautifulSoup

sys.path.insert(0, str(Path(__file__).parent))
from config import CSV_PATHS, SCREENER_BASE_URL, STOCK_TICKERS, TICKER_COLUMN_NAMES, TRENDLYNE_IDS

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
        "AppleWebKit/537.36 (KHTML, like Gecko) "
        "Chrome/124.0.0.0 Safari/537.36"
    ),
}

SESSION = requests.Session()
SESSION.headers.update(HEADERS)


# ── Helpers ───────────────────────────────────────────────────────────────────

def _format_date(d: date) -> str:
    return d.strftime("%d/%m/%y")


def _last_date_in_sheet(ws, check_data_col=None):
    """Return last date in column A where data column is also filled, or None.

    check_data_col: 1-based column index that must be non-empty for the row to count.
    Use this to skip pre-populated date rows that have no actual data.
    """
    import datetime as dt
    n_cols = max(1, check_data_col) if check_data_col else 1
    last = None
    for row in ws.iter_rows(min_row=2, max_col=n_cols, values_only=True):
        val = row[0]
        if val is None:
            continue
        # Skip rows where the data column is empty
        if check_data_col is not None:
            data_val = row[check_data_col - 1] if len(row) >= check_data_col else None
            if data_val in (None, ""):
                continue
        if isinstance(val, dt.datetime):
            last = val.date()
        elif isinstance(val, dt.date):
            last = val
        elif isinstance(val, str):
            try:
                parts = val.strip().split("/")
                d, m, y = int(parts[0]), int(parts[1]), int(parts[2])
                if y < 100:
                    y += 2000
                last = date(y, m, d)
            except Exception:
                pass
    return last


def _get_column_index(ws, ticker: str):
    """Find the 1-based column index of ticker in row 1, or None."""
    for col in range(1, ws.max_column + 2):
        val = ws.cell(row=1, column=col).value
        if val is None:
            return None
        if str(val).strip().upper() == ticker.upper():
            return col
    return None


def _ensure_ticker_column(ws, ticker: str) -> int:
    """Return existing column for ticker, or add it and return new index."""
    idx = _get_column_index(ws, ticker)
    if idx is not None:
        return idx
    new_col = ws.max_column + 1
    ws.cell(row=1, column=new_col, value=ticker)
    return new_col


# ── Scraping ──────────────────────────────────────────────────────────────────

def _scrape_screener(ticker: str):
    """
    Scrape current price and P/B ratio for a ticker from Screener.in.
    Returns {"price": float, "pb": float} or None.
    """
    base_url = SCREENER_BASE_URL.format(ticker=ticker)
    consolidated_url = base_url.rstrip("/") + "/consolidated/"
    try:
        r = SESSION.get(consolidated_url, timeout=20)
        if r.status_code == 404:
            # Fallback to standalone if no consolidated page
            r = SESSION.get(base_url, timeout=20)
        r.raise_for_status()
    except Exception as e:
        print(f"[STOCKS] Failed to fetch {ticker}: {e}")
        return None

    soup = BeautifulSoup(r.text, "lxml")

    price = None
    pb = None

    # ── Current price ─────────────────────────────────────────────────────
    # Screener shows price in <span class="number"> inside the stock header
    price_candidates = [
        soup.find("span", {"id": "company-current-price"}),
        soup.find("div", class_=re.compile(r"stock-price|current-price")),
    ]
    for el in price_candidates:
        if el:
            text = re.sub(r"[^\d.]", "", el.get_text())
            try:
                price = float(text)
                break
            except Exception:
                pass

    # Fallback: look for "Current Price" label in the top ratios section
    if price is None:
        for li in soup.find_all("li"):
            name = li.find("span", class_="name")
            number = li.find("span", class_="number")
            if name and number:
                if "current price" in name.get_text(strip=True).lower():
                    text = re.sub(r"[^\d.]", "", number.get_text())
                    try:
                        price = float(text)
                    except Exception:
                        pass

    # ── P/B Ratio ─────────────────────────────────────────────────────────
    # Screener shows "Book Value" (per share), not always "Price to Book".
    # If only book value is found, compute P/B = price / book_value.
    pb_direct_kws = ["price to book", "p/b", "pb ratio"]
    book_value = None

    for li in soup.find_all("li"):
        name_el = li.find("span", class_="name")
        num_el  = li.find("span", class_="number")
        if not (name_el and num_el):
            continue
        label = name_el.get_text(strip=True).lower()
        text  = re.sub(r"[^\d.]", "", num_el.get_text())
        try:
            val = float(text)
        except Exception:
            continue
        if any(kw in label for kw in pb_direct_kws):
            pb = val
            break                          # direct P/B found — done
        if "book value" in label and book_value is None:
            book_value = val               # per-share; keep searching for direct P/B

    # Compute P/B from book value if no direct ratio found
    if pb is None and book_value is not None and price is not None and book_value > 0:
        pb = round(price / book_value, 2)

    # Fallback: raw text search for "Price to Book Value"
    if pb is None:
        pb_match = re.search(
            r"Price\s+to\s+Book\s+[Vv]alue\s*[\n\r\s]*(\d+\.?\d*)",
            soup.get_text(" ", strip=True)
        )
        if pb_match:
            try:
                pb = float(pb_match.group(1))
            except Exception:
                pass

    if price is None and pb is None:
        print(f"[STOCKS] WARNING: No data found for {ticker} at {url}")
        return None

    return {"price": price, "pb": pb}


def _scrape_nse_price(ticker: str):
    """Fetch current price from NSE public API as fallback when Screener.in fails."""
    import urllib.request
    import json as _json
    url = f"https://www.nseindia.com/api/quote-equity?symbol={ticker}"
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
        ),
        "Accept": "application/json",
        "Referer": "https://www.nseindia.com/",
    }
    try:
        req = urllib.request.Request(url, headers=headers)
        with urllib.request.urlopen(req, timeout=15) as resp:
            data = _json.loads(resp.read().decode("utf-8"))
        price = data.get("priceInfo", {}).get("lastPrice")
        return float(price) if price is not None else None
    except Exception as e:
        print(f"[STOCKS] NSE fallback failed for {ticker}: {e}")
        return None


def _fetch_trendlyne_pb(ticker: str) -> float | None:
    """Fetch latest P/B ratio from Trendlyne API (consolidated book value)."""
    tl_id = TRENDLYNE_IDS.get(ticker)
    if not tl_id:
        return None
    try:
        r = SESSION.get(
            f"https://trendlyne.com/mapp/v1/stock/chart-data/{tl_id}/PBV_A_SHARE_NOW/",
            timeout=10,
        )
        pts = r.json().get("body", {}).get("eodData", [])
        if pts:
            return round(pts[0][1], 2)  # pts[0] = latest [timestamp_ms, pb_value]
    except Exception as e:
        print(f"[STOCKS] Trendlyne P/B fetch failed for {ticker}: {e}")
    return None


# ── XLSX update ───────────────────────────────────────────────────────────────

def _update_xlsx(results, target_date):
    """
    Append one row of prices and one row of P/B ratios to stock.xlsx.
    results: {ticker: {"price": float, "pb": float} or None}
    """
    import openpyxl

    xlsx_path = CSV_PATHS["stocks"]
    p = Path(xlsx_path)

    if p.exists():
        wb = openpyxl.load_workbook(xlsx_path)
    else:
        wb = openpyxl.Workbook()
        wb.active.title = "Prices"
        wb.create_sheet("P/B")

    prices_ws = wb["Stock Prices"]
    pb_ws     = wb["Stock P-B"]

    # Headers already exist in "Stock Prices" and "Stock P-B"

    # Independent idempotency checks — so a missed P/B write is retried even
    # when prices are already present (and vice-versa).
    last_price_date = _last_date_in_sheet(prices_ws, check_data_col=2)
    skip_prices = bool(last_price_date and last_price_date >= target_date)

    last_pb_date = _last_date_in_sheet(pb_ws, check_data_col=2)
    skip_pb = bool(last_pb_date and last_pb_date >= target_date)

    if skip_prices and skip_pb:
        print(f"[STOCKS] Data for {target_date} already in both sheets. Skipping.")
        return True
    if skip_prices:
        print(f"[STOCKS] Prices for {target_date} already present. Writing P/B only.")
    if skip_pb:
        print(f"[STOCKS] P/B for {target_date} already present. Writing prices only.")

    date_str = _format_date(target_date)

    price_new_row = prices_ws.max_row + 1
    pb_new_row    = pb_ws.max_row + 1

    if not skip_prices:
        prices_ws.cell(price_new_row, 1, date_str)
    if not skip_pb:
        pb_ws.cell(pb_new_row, 1, date_str)

    for ticker, data in results.items():
        if data is None:
            continue
        col_name  = TICKER_COLUMN_NAMES.get(ticker, ticker)

        if not skip_prices and data.get("price") is not None:
            price_col = _ensure_ticker_column(prices_ws, col_name)
            prices_ws.cell(price_new_row, price_col, data["price"])
        if not skip_pb and data.get("pb") is not None:
            pb_col = _ensure_ticker_column(pb_ws, col_name)
            pb_ws.cell(pb_new_row, pb_col, data["pb"])

    wb.save(xlsx_path)
    return True


# ── Main logic ────────────────────────────────────────────────────────────────

_IST = timezone(timedelta(hours=5, minutes=30))


def scrape_stocks(target_date=None):
    if target_date is None:
        target_date = datetime.now(_IST).date() - timedelta(days=1)

    print(f"[STOCKS] Scraping {len(STOCK_TICKERS)} tickers for {target_date}...")

    results: dict[str, dict | None] = {}
    for ticker in STOCK_TICKERS:
        print(f"[STOCKS] Fetching {ticker}...")
        data = _scrape_screener(ticker)
        # NSE fallback: if Screener.in returned no price, try NSE public API
        if data is None or data.get("price") is None:
            print(f"[STOCKS] {ticker}: Screener price unavailable, trying NSE fallback...")
            nse_price = _scrape_nse_price(ticker)
            if nse_price is not None:
                data = data or {}
                data["price"] = nse_price
        results[ticker] = data if data else None
        if data:
            print(f"[STOCKS] {ticker}: price={data.get('price')}, P/B={data.get('pb')}")
        time.sleep(2)  # Be polite to Screener.in

    if all(v is None for v in results.values()):
        print("[STOCKS] ERROR: All tickers failed.")
        return False

    success = _update_xlsx(results, target_date)
    if success:
        print(f"[STOCKS] Done for {target_date}.")
    return success


if __name__ == "__main__":
    ok = scrape_stocks()
    sys.exit(0 if ok else 1)
