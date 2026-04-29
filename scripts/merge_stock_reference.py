"""
One-time merge: overwrite Jan 1–Apr 28 2026 price + P/B rows in stock.xlsx
with authoritative consolidated data from the reference file.
"""

import sys
from datetime import date, datetime
from pathlib import Path

import openpyxl
import pandas as pd

sys.path.insert(0, str(Path(__file__).parent))
from config import CSV_PATHS

SOURCE_PATH = (
    Path(__file__).parent.parent.parent
    / "Power Stocks - 1st jan-28th april 2026.xlsx"
)
START = date(2026, 1, 1)
END   = date(2026, 4, 28)


def _parse_date(val):
    if val is None:
        return None
    if isinstance(val, datetime):
        return val.date()
    if isinstance(val, date):
        return val
    if isinstance(val, str):
        try:
            parts = val.strip().split("/")
            dd, mm, yy = int(parts[0]), int(parts[1]), int(parts[2])
            if yy < 100:
                yy += 2000
            return date(yy, mm, dd)
        except Exception:
            pass
    return None


def _fmt(d: date) -> str:
    return d.strftime("%d/%m/%y")


def _normalize(name: str) -> str:
    return str(name).strip().rstrip(".")


def main():
    # ── 1. Load source file ────────────────────────────────────────────────────
    src = pd.read_excel(SOURCE_PATH)
    src = src.dropna(subset=["Company Name", "NDP_Date"])
    src["_date"] = pd.to_datetime(src["NDP_Date"]).dt.date
    src["_norm"] = src["Company Name"].apply(_normalize)

    # Build {date: {norm_company: (price, pb)}}
    src_data: dict[date, dict[str, tuple]] = {}
    for _, row in src.iterrows():
        d = row["_date"]
        if START <= d <= END:
            src_data.setdefault(d, {})[row["_norm"]] = (
                float(row["NDP_Close"]),
                float(row["NDP_Consolidated Price BV"]),
            )

    print(f"Source: {len(src_data)} unique trading dates, {len(src)} rows")

    # ── 2. Load stock.xlsx ─────────────────────────────────────────────────────
    xlsx_path = CSV_PATHS["stocks"]
    wb = openpyxl.load_workbook(xlsx_path)
    pr_ws = wb["Stock Prices"]
    pb_ws = wb["Stock P-B"]

    # Build header → col index maps (1-based), normalizing names
    def col_map(ws):
        m = {}
        for c in range(1, ws.max_column + 1):
            h = ws.cell(1, c).value
            if h:
                m[_normalize(str(h))] = c
        return m

    pr_cols = col_map(pr_ws)
    pb_cols = col_map(pb_ws)

    # Verify all 11 companies resolve
    missing = [n for n in src["_norm"].unique() if n not in pr_cols]
    if missing:
        print(f"WARNING: companies not found in xlsx headers: {missing}")

    # ── 3. Scan existing rows in range ─────────────────────────────────────────
    def scan_rows(ws):
        row_map = {}  # date -> row_idx
        for ri in range(2, ws.max_row + 1):
            d = _parse_date(ws.cell(ri, 1).value)
            if d and START <= d <= END:
                row_map[d] = ri
        return row_map

    pr_row_map = scan_rows(pr_ws)
    pb_row_map = scan_rows(pb_ws)

    print(f"Stock Prices: {len(pr_row_map)} rows in range")
    print(f"Stock P-B:   {len(pb_row_map)} rows in range")

    # ── 4. Pass A — overwrite matching rows; collect stale rows ───────────────
    def overwrite_rows(ws, row_map, cols, value_key):
        written, stale = 0, []
        for d, ri in row_map.items():
            if d not in src_data:
                stale.append(ri)
                print(f"  Stale row {ri}: {d} ({d.strftime('%A')}) — not in reference, will delete")
                continue
            day_data = src_data[d]
            for norm_company, (price, pb) in day_data.items():
                col = cols.get(norm_company)
                if col is None:
                    continue
                val = price if value_key == "price" else pb
                ws.cell(ri, col, round(val, 4))
            written += 1
        return written, stale

    pr_written, pr_stale = overwrite_rows(pr_ws, pr_row_map, pr_cols, "price")
    pb_written, pb_stale = overwrite_rows(pb_ws, pb_row_map, pb_cols, "pb")

    # ── 5. Pass B — delete stale rows (reverse order) ─────────────────────────
    for ri in sorted(pr_stale, reverse=True):
        pr_ws.delete_rows(ri)
    for ri in sorted(pb_stale, reverse=True):
        pb_ws.delete_rows(ri)

    print(f"Deleted {len(pr_stale)} stale rows from Prices, {len(pb_stale)} from P-B")

    # ── 6. Pass C — insert missing dates ──────────────────────────────────────
    existing_pr = set(scan_rows(pr_ws))  # re-scan after deletions
    existing_pb = set(scan_rows(pb_ws))

    inserted = 0
    for d in sorted(src_data):
        price_missing = d not in existing_pr
        pb_missing    = d not in existing_pb
        if not price_missing and not pb_missing:
            continue

        day_data = src_data[d]
        date_str = _fmt(d)

        if price_missing:
            ri = pr_ws.max_row + 1
            pr_ws.cell(ri, 1, date_str)
            for norm_company, (price, pb) in day_data.items():
                col = pr_cols.get(norm_company)
                if col:
                    pr_ws.cell(ri, col, round(price, 4))
            print(f"  Inserted Prices row {ri}: {d}")

        if pb_missing:
            ri = pb_ws.max_row + 1
            pb_ws.cell(ri, 1, date_str)
            for norm_company, (price, pb) in day_data.items():
                col = pb_cols.get(norm_company)
                if col:
                    pb_ws.cell(ri, col, round(pb, 4))
            print(f"  Inserted P-B row {ri}: {d}")

        inserted += 1

    # ── 7. Save ────────────────────────────────────────────────────────────────
    wb.save(xlsx_path)

    print(f"\nDone.")
    print(f"  Prices : {pr_written} overwritten, {len(pr_stale)} deleted, {inserted} inserted")
    print(f"  P-B    : {pb_written} overwritten, {len(pb_stale)} deleted, {inserted} inserted")

    # ── 8. Spot-check verification ─────────────────────────────────────────────
    wb2 = openpyxl.load_workbook(xlsx_path)
    pb2 = wb2["Stock P-B"]
    pr2 = wb2["Stock Prices"]

    def find_row(ws, target_date):
        for ri in range(2, ws.max_row + 1):
            if _parse_date(ws.cell(ri, 1).value) == target_date:
                return ri
        return None

    def col_idx(ws, norm_name):
        for c in range(1, ws.max_column + 1):
            h = ws.cell(1, c).value
            if h and _normalize(str(h)) == norm_name:
                return c
        return None

    checks = [
        ("Stock P-B",    pb2, date(2026, 1, 2), "NTPC Ltd",              1.74),
        ("Stock P-B",    pb2, date(2026, 1, 2), "Adani Green Energy Ltd", 8.81),
        ("Stock Prices", pr2, date(2026, 4, 28), "NTPC Ltd",             408.0),
    ]
    print("\nSpot checks:")
    for sheet, ws, d, company, expected in checks:
        ri  = find_row(ws, d)
        ci  = col_idx(ws, _normalize(company))
        val = ws.cell(ri, ci).value if (ri and ci) else "NOT FOUND"
        ok  = isinstance(val, (int, float)) and abs(val - expected) / expected < 0.05
        print(f"  {sheet} | {d} | {company}: {val} (expected ~{expected}) {'✓' if ok else '✗ CHECK'}")


if __name__ == "__main__":
    main()
