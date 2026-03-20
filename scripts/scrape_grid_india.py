"""
Scrape Grid India Daily PSP Report (Excel) and append to:
  - supply.csv        : Total energy met (MU)      — MOP_E!H8  (row: "Energy Met (MU)")
  - Peak Demand.csv   : Peak demand met (GW)        — MOP_E!H13 (row: "Maximum Demand Met…"), /1000
  - generation.csv    : Total generation, Thermal (derived), Renewable — MOP_E section G

The Excel file is published on grid-india.in for the *previous* day.
The DATA date is embedded in the Excel filename (not the upload/listing date).

Run pattern: idempotent, exits 0 on success OR if data already present, exits 2 if Excel
not yet published (GitHub Actions workflow retries every 30 min from 9am–11am IST).

Usage:
  python scrape_grid_india.py                  # yesterday (IST)
  python scrape_grid_india.py --date 2026-03-17
"""

import argparse
import csv
import io
import re
import sys
from datetime import date, datetime, timedelta, timezone
from pathlib import Path

import requests

sys.path.insert(0, str(Path(__file__).parent))
from config import CSV_PATHS, GRID_INDIA_REPORTS_URL, GRID_INDIA_PDF_BASE

IST = timezone(timedelta(hours=5, minutes=30))

HEADERS = {
    "User-Agent": (
        "Mozilla/5.0 (Windows NT 10.0; Win64; x64) AppleWebKit/537.36 "
        "(KHTML, like Gecko) Chrome/124.0.0.0 Safari/537.36"
    ),
    "Accept": "application/json, text/plain, */*",
    "Accept-Language": "en-IN,en;q=0.9",
    "Content-Type": "application/json",
    "Origin": "https://grid-india.in",
    "Referer": "https://grid-india.in/en/reports/daily-psp-report",
}


# ── CSV helpers ────────────────────────────────────────────────────────────────

def _last_date_in_csv(path: str):
    p = Path(path)
    if not p.exists():
        return None
    with open(p, newline="") as f:
        rows = list(csv.reader(f))
    data_rows = [r for r in rows if r and not r[0].strip().lower().startswith("date")]
    if not data_rows:
        return None
    last = data_rows[-1][0].strip()
    try:
        d, m, y = last.split("/")
        y = int(y)
        if y < 100:
            y += 2000
        return date(y, int(m), int(d))
    except Exception:
        return None


def _fy_label(d: date) -> str:
    """Return Indian financial year label for a date, e.g. '2025-26'."""
    if d.month >= 4:
        return f"{d.year}-{str(d.year + 1)[2:]}"
    return f"{d.year - 1}-{str(d.year)[2:]}"


def _fy_labels_in_range(start: date, end: date) -> list:
    """Return sorted list of FY labels that overlap [start, end]."""
    labels = set()
    d = start
    while d <= end:
        lbl = _fy_label(d)
        labels.add(lbl)
        y = int(lbl.split("-")[0])
        d = date(y + 1, 4, 1)
    labels.add(_fy_label(end))
    return sorted(labels)


def _format_date(d: date) -> str:
    return d.strftime("%d/%m/%y")


def _append_csv(path: str, row: list) -> None:
    p = Path(path)
    # Ensure file ends with a newline before appending
    if p.exists() and p.stat().st_size > 0:
        with open(p, "rb") as f:
            f.seek(-1, 2)
            last_byte = f.read(1)
        if last_byte not in (b"\n", b"\r"):
            with open(p, "a") as f:
                f.write("\n")
    with open(p, "a", newline="") as f:
        csv.writer(f).writerow(row)


# ── Excel URL discovery via Grid India REST API ────────────────────────────────
# Grid India exposes an undocumented REST API at webapi.grid-india.in that returns
# file listings. The CDN (webcdn.grid-india.in) serves files directly without any
# browser or IP restrictions — no Playwright needed.

_GRID_API_FILE_URL = "https://webapi.grid-india.in/api/v1/file"
_GRID_CDN_BASE     = "https://webcdn.grid-india.in"


def _find_excel_url(target_date: date):
    """
    Use Grid India REST API to find the XLS download URL for target_date.
    No browser required — works from any IP including GitHub Actions.
    """
    fy           = _fy_label(target_date)           # e.g. "2025-26"
    month        = target_date.strftime("%m")        # e.g. "03"
    date_prefix  = target_date.strftime("%d.%m.%y") # e.g. "18.03.26"

    print(f"[GRID] Querying Grid India API for {target_date} (FY={fy}, month={month})...")
    items = None
    for attempt in range(1, 4):  # up to 3 attempts
        try:
            resp = requests.post(
                _GRID_API_FILE_URL,
                json={"_source": "GRDW", "_type": "DAILY_PSP_REPORT",
                      "_fileDate": fy, "_month": month},
                headers=HEADERS,
                timeout=30,
            )
            resp.raise_for_status()
            items = resp.json().get("retData") or []
            break
        except Exception as e:
            print(f"[GRID] API error (attempt {attempt}/3): {e}")
            if attempt < 3:
                import time as _time
                _time.sleep(5 * attempt)  # 5s, 10s back-off
    if items is None:
        return None

    for item in items:
        fp = item.get("FilePath", "")
        if date_prefix in fp and fp.endswith(".xls"):
            url = f"{_GRID_CDN_BASE}/{fp}"
            print(f"[GRID] Found XLS: {url}")
            return url

    print(f"[GRID] No XLS found for {date_prefix} in API response ({len(items)} files).")
    return None


def _collect_all_excel_urls(start_date: date, end_date: date) -> dict:
    """
    Use Grid India REST API to collect {date: url} for all XLS files in [start_date, end_date].
    No browser required — works from any IP including GitHub Actions.
    """
    import re as _re

    # Build the set of (fy, month) pairs that overlap the date range
    fy_months = set()
    d = start_date.replace(day=1)
    while d <= end_date:
        fy_months.add((_fy_label(d), d.strftime("%m")))
        if d.month == 12:
            d = d.replace(year=d.year + 1, month=1)
        else:
            d = d.replace(month=d.month + 1)
    fy_months.add((_fy_label(end_date), end_date.strftime("%m")))

    print(f"[GRID-API] Collecting XLS URLs {start_date} → {end_date} "
          f"({len(fy_months)} FY-month buckets)...")

    url_map = {}
    for fy, month in sorted(fy_months):
        print(f"[GRID-API] Fetching FY={fy} month={month}...")
        try:
            resp = requests.post(
                _GRID_API_FILE_URL,
                json={"_source": "GRDW", "_type": "DAILY_PSP_REPORT",
                      "_fileDate": fy, "_month": month},
                headers=HEADERS,
                timeout=30,
            )
            resp.raise_for_status()
            items = resp.json().get("retData") or []
        except Exception as e:
            print(f"[GRID-API] Error for FY={fy} month={month}: {e}")
            continue

        for item in items:
            fp = item.get("FilePath", "")
            if not fp.endswith(".xls"):
                continue
            title = item.get("Title_", "")
            m = _re.match(r"(\d{2})\.(\d{2})\.(\d{2})_", title)
            if m:
                day_, mon_, yr_ = int(m.group(1)), int(m.group(2)), 2000 + int(m.group(3))
                try:
                    file_date = date(yr_, mon_, day_)
                    if start_date <= file_date <= end_date:
                        url_map[file_date] = f"{_GRID_CDN_BASE}/{fp}"
                except ValueError:
                    pass

        print(f"[GRID-API] FY={fy} month={month}: total so far = {len(url_map)}")

    print(f"[GRID-API] Total XLS URLs collected: {len(url_map)}")
    return url_map


# ── Excel loading (supports both .xls and .xlsx) ──────────────────────────────

class _Cell:
    """Minimal cell wrapper so xlrd and openpyxl share the same interface."""
    __slots__ = ("value",)
    def __init__(self, value):
        self.value = value if value != "" else None


class _SheetAdapter:
    """
    Wraps either an openpyxl worksheet or an xlrd sheet so that
    callers can use .cell(row, column) with 1-based indices and .max_row.
    """
    def __init__(self, sheet, xlrd_mode: bool = False):
        self._s = sheet
        self._xlrd = xlrd_mode

    @property
    def max_row(self):
        return self._s.nrows if self._xlrd else self._s.max_row

    def cell(self, row: int, column: int):
        if self._xlrd:
            try:
                val = self._s.cell_value(row - 1, column - 1)
            except IndexError:
                val = None
            return _Cell(val)
        return self._s.cell(row=row, column=column)


def _open_sheet(excel_bytes: bytes, sheet_name: str):
    """
    Open an Excel file (either .xls or .xlsx) from bytes and return a
    _SheetAdapter for the named sheet, or None on failure.
    """
    # Try openpyxl first (xlsx)
    try:
        from openpyxl import load_workbook
        wb = load_workbook(io.BytesIO(excel_bytes), data_only=True)
        if sheet_name in wb.sheetnames:
            return _SheetAdapter(wb[sheet_name], xlrd_mode=False)
        print(f"[GRID] openpyxl: sheet '{sheet_name}' not found. Available: {wb.sheetnames}")
    except Exception:
        pass

    # Fall back to xlrd (xls)
    try:
        import xlrd
        wb = xlrd.open_workbook(file_contents=excel_bytes)
        ws = wb.sheet_by_name(sheet_name)
        return _SheetAdapter(ws, xlrd_mode=True)
    except Exception as e:
        print(f"[GRID] xlrd failed: {e}")

    return None


# ── Excel parsing helpers ──────────────────────────────────────────────────────

def _find_row_in_col(ws: _SheetAdapter, keyword: str, col: int = 1, start: int = 1, max_scan: int = 200):
    """
    Like _find_row but searches an arbitrary column (1-indexed).
    Useful for sections whose labels are not in column A.
    """
    kw = keyword.lower()
    end = min(start + max_scan, ws.max_row + 1)
    for r in range(start, end):
        val = str(ws.cell(row=r, column=col).value or "").strip().lower()
        if kw in val:
            return r
    return None


def _excel_time_to_str(val) -> str:
    """Convert an Excel time value to HH:MM string.
    xlrd returns time as a float (fraction of day); openpyxl returns datetime.time.
    Strings and None are handled gracefully.
    """
    if val is None:
        return ""
    # datetime.time from openpyxl
    try:
        import datetime as _dt
        if isinstance(val, _dt.time):
            return val.strftime("%H:%M")
    except Exception:
        pass
    # float from xlrd: fraction of a 24-hour day
    if isinstance(val, float) and 0.0 <= val <= 1.0:
        total_min = int(round(val * 24 * 60))
        h, m = divmod(total_min, 60)
        return f"{h:02d}:{m:02d}"
    return str(val).strip()


def _find_row(ws: _SheetAdapter, keyword: str, start: int = 1, max_scan: int = 200):
    """
    Scan column A from start to start+max_scan for a cell whose string value
    contains keyword (case-insensitive). Returns the row number or None.
    """
    kw = keyword.lower()
    end = min(start + max_scan, ws.max_row + 1)
    for r in range(start, end):
        val = str(ws.cell(row=r, column=1).value or "").strip().lower()
        if kw in val:
            return r
    return None


def _find_row_after(ws: _SheetAdapter, keyword: str, start: int, max_scan: int = 20):
    """Like _find_row but starts at start+1 (for anchored section searches)."""
    return _find_row(ws, keyword, start=start + 1, max_scan=max_scan)


def _cell_float(ws: _SheetAdapter, row: int, col: int):
    """Return float value of cell or None."""
    if row is None:
        return None
    try:
        v = ws.cell(row=row, column=col).value
        if v is None:
            return None
        return float(str(v).replace(",", "").strip())
    except (ValueError, TypeError):
        return None


def _parse_mop_e(ws: _SheetAdapter) -> dict:
    """
    Parse the MOP_E sheet and return a dict with keys:
      supply_mu, peak_demand_gw, total_gen_mu, renewable_mu, thermal_mu,
      solar_peak_gw, solar_peak_time, nonsolar_peak_gw, nonsolar_peak_time
    Missing values are absent from the dict.
    """
    result = {}
    H = 8  # column index for "All India" / Total column

    # ── Supply (Energy Met MU) ────────────────────────────────────────────────
    r = _find_row(ws, "energy met")
    if r:
        v = _cell_float(ws, r, H)
        if v is not None:
            result["supply_mu"] = round(v, 0)
            print(f"[GRID] Supply (row {r}): {result['supply_mu']} MU")
        else:
            print(f"[GRID] WARNING: 'Energy Met' found at row {r} but H{r} is empty/non-numeric")
    else:
        print("[GRID] WARNING: 'Energy Met' header not found in MOP_E column A")

    # ── Peak Demand (MW → GW) ─────────────────────────────────────────────────
    r = _find_row(ws, "maximum demand met")
    if r:
        v = _cell_float(ws, r, H)
        if v is not None:
            result["peak_demand_gw"] = round(v / 1000, 2)
            print(f"[GRID] Peak Demand (row {r}): {v} MW → {result['peak_demand_gw']} GW")
        else:
            print(f"[GRID] WARNING: 'Maximum Demand Met' found at row {r} but H{r} is empty")
    else:
        print("[GRID] WARNING: 'Maximum Demand Met' header not found in MOP_E column A")

    # ── Total Generation ──────────────────────────────────────────────────────
    # Anchored: find the section header first, then "Total" within 20 rows
    section_row = _find_row(ws, "G. Sourcewise generation (Gross)")
    if section_row is None:
        section_row = _find_row(ws, "Sourcewise generation")  # fallback partial
    if section_row:
        total_row = _find_row_after(ws, "Total", start=section_row, max_scan=20)
        if total_row:
            v = _cell_float(ws, total_row, H)
            if v is not None:
                result["total_gen_mu"] = round(v, 0)
                print(f"[GRID] Total Generation (row {total_row}): {result['total_gen_mu']} MU")
            else:
                print(f"[GRID] WARNING: Total row {total_row} found but H value is empty")
        else:
            print(f"[GRID] WARNING: 'Total' not found within 20 rows of section row {section_row}")
    else:
        print("[GRID] WARNING: 'G. Sourcewise generation' section not found")

    # ── Renewable (RES) ───────────────────────────────────────────────────────
    r = _find_row(ws, "RES (Wind, Solar, Biomass")
    if r:
        v = _cell_float(ws, r, H)
        if v is not None:
            result["renewable_mu"] = round(v, 0)
            print(f"[GRID] Renewable (row {r}): {result['renewable_mu']} MU")
        else:
            print(f"[GRID] WARNING: RES row {r} found but H value is empty")
    else:
        print("[GRID] WARNING: 'RES (Wind, Solar, Biomass' row not found in MOP_E")

    # ── Thermal = Total − Renewable (derived, balancing figure) ──────────────
    if "total_gen_mu" in result and "renewable_mu" in result:
        result["thermal_mu"] = round(result["total_gen_mu"] - result["renewable_mu"], 0)
        print(f"[GRID] Thermal (derived): {result['thermal_mu']} MU")

    # ── Solar / Non-Solar Hour Peak Demand ────────────────────────────────────
    # Section header is in col E (≈E92): "All India Peak Demand and shortage at Solar and Non-Solar Hour"
    # Solar row ≈ E94 (keyword "solar hr"), Non-Solar row ≈ E95 (keyword "non solar hr")
    # F col = Max Demand Met (MW), H col = time of peak
    E = 5   # column E
    F = 6   # column F — Max Demand Met (MW)
    TH = 8  # column H — time of peak (reuse, distinct from H above)
    section_row = _find_row_in_col(ws, "all india peak demand and shortage", col=E, start=80, max_scan=30)
    if section_row:
        solar_row    = _find_row_in_col(ws, "solar hr",      col=E, start=section_row + 1, max_scan=8)
        nonsolar_row = _find_row_in_col(ws, "non-solar hr",  col=E, start=section_row + 1, max_scan=8)

        if solar_row:
            v = _cell_float(ws, solar_row, F)
            if v is not None:
                result["solar_peak_gw"] = round(v / 1000, 2)
            result["solar_peak_time"] = _excel_time_to_str(ws.cell(row=solar_row, column=TH).value)
            print(f"[GRID] Solar Peak (row {solar_row}): {result.get('solar_peak_gw')} GW @ {result.get('solar_peak_time')}")
        else:
            print("[GRID] WARNING: 'solar hr' row not found in col E")

        if nonsolar_row:
            v = _cell_float(ws, nonsolar_row, F)
            if v is not None:
                result["nonsolar_peak_gw"] = round(v / 1000, 2)
            result["nonsolar_peak_time"] = _excel_time_to_str(ws.cell(row=nonsolar_row, column=TH).value)
            print(f"[GRID] Non-Solar Peak (row {nonsolar_row}): {result.get('nonsolar_peak_gw')} GW @ {result.get('nonsolar_peak_time')}")
        else:
            print("[GRID] WARNING: 'non solar hr' row not found in col E")
    else:
        print("[GRID] WARNING: Solar/Non-Solar peak section not found in MOP_E col E")

    return result


# ── Main ──────────────────────────────────────────────────────────────────────

def scrape_grid_india(target_date=None, excel_url=None, force=False) -> bool:
    """
    Returns True on success. Raises SystemExit(2) if Excel not yet published.

    excel_url: optional pre-fetched URL (skips Playwright discovery).
    force: skip the idempotency check (use during backfill for gap-filling).
    """
    if target_date is None:
        target_date = datetime.now(IST).date() - timedelta(days=1)

    if not force:
        # Idempotency: check supply.csv as representative; all CSVs are written together
        last = _last_date_in_csv(CSV_PATHS["supply"])
        if last and last >= target_date:
            # Also check solar-nonsolar separately (may lag if newly added)
            last_sn = _last_date_in_csv(CSV_PATHS["demand_solar_nonsolar"])
            if last_sn and last_sn >= target_date:
                print(f"[GRID] Data for {target_date} already present in all CSVs. Skipping.")
                return True
            # Supply already written but solar-nonsolar missing — proceed to (re-)extract
            print(f"[GRID] Supply up to date but solar-nonsolar CSV needs {target_date}. Proceeding.")


    print(f"[GRID] Looking for Excel file for {target_date}...")

    if excel_url is None:
        excel_url = _find_excel_url(target_date)
    if excel_url is None:
        print(f"[GRID] Excel not yet available for {target_date}. Will retry.")
        sys.exit(2)  # GitHub Actions retry signal

    print(f"[GRID] Downloading: {excel_url}")
    try:
        r = requests.get(excel_url, headers=HEADERS, timeout=60)
        r.raise_for_status()
        excel_bytes = r.content
    except Exception as e:
        print(f"[GRID] Failed to download Excel: {e}")
        return False

    ws = _open_sheet(excel_bytes, "MOP_E")
    if ws is None:
        print("[GRID] ERROR: Could not open MOP_E sheet from workbook.")
        return False

    data = _parse_mop_e(ws)
    if not data:
        print("[GRID] ERROR: Could not extract any data from MOP_E sheet.")
        return False

    date_str = _format_date(target_date)
    wrote_any = False

    # Write supply.csv
    if "supply_mu" in data:
        _append_csv(CSV_PATHS["supply"], [date_str, data["supply_mu"]])
        wrote_any = True
    else:
        print("[GRID] WARNING: supply_mu missing — not writing to supply.csv")

    # Write Peak Demand.csv
    if "peak_demand_gw" in data:
        _append_csv(CSV_PATHS["demand"], [date_str, data["peak_demand_gw"]])
        wrote_any = True
    else:
        print("[GRID] WARNING: peak_demand_gw missing — not writing to Peak Demand.csv")

    # Write generation.csv (Date, Total, Coal/Thermal, Renewable)
    if "total_gen_mu" in data:
        thermal  = data.get("thermal_mu", "")
        renewabl = data.get("renewable_mu", "")
        _append_csv(CSV_PATHS["generation"], [date_str, data["total_gen_mu"], thermal, renewabl])
        wrote_any = True
    else:
        print("[GRID] WARNING: total_gen_mu missing — not writing to generation.csv")

    # Write Peak Demand Solar-NonSolar.csv
    if "solar_peak_gw" in data or "nonsolar_peak_gw" in data:
        sn_path = CSV_PATHS["demand_solar_nonsolar"]
        if not Path(sn_path).exists():
            _append_csv(sn_path, ["date", "solar_gw", "solar_time", "nonsolar_gw", "nonsolar_time"])
        _append_csv(sn_path, [
            date_str,
            data.get("solar_peak_gw", ""),
            data.get("solar_peak_time", ""),
            data.get("nonsolar_peak_gw", ""),
            data.get("nonsolar_peak_time", ""),
        ])
        wrote_any = True
        print(f"[GRID] Solar-NonSolar Peak: solar={data.get('solar_peak_gw')} GW, "
              f"nonsolar={data.get('nonsolar_peak_gw')} GW")
    else:
        print("[GRID] WARNING: solar/nonsolar peak missing — not writing to Peak Demand Solar-NonSolar.csv")

    if wrote_any:
        print(f"[GRID] Done for {target_date}.")
        return True

    print("[GRID] ERROR: No CSV rows written.")
    return False


if __name__ == "__main__":
    parser = argparse.ArgumentParser()
    parser.add_argument("--date", help="Target date YYYY-MM-DD (default: yesterday IST)")
    args = parser.parse_args()

    target = None
    if args.date:
        target = date.fromisoformat(args.date)

    success = scrape_grid_india(target)
    sys.exit(0 if success else 1)
